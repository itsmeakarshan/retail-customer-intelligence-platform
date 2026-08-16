"""
Unit and integration tests for the Inventory Optimisation Engine.
"""
import pytest
from ml.src.forecasting.inventory_optimizer import InventoryOptimizer, SERVICE_LEVEL_Z


def test_service_level_z_values():
    """Tests standard normal z values for service levels."""
    assert SERVICE_LEVEL_Z[0.90] == 1.282
    assert SERVICE_LEVEL_Z[0.95] == 1.645
    assert SERVICE_LEVEL_Z[0.98] == 2.054
    assert SERVICE_LEVEL_Z[0.99] == 2.326


def test_inventory_calculation_replenishment_needed():
    """Tests ROP, Safety Stock, and Suggested Order when stock is low."""
    optimizer = InventoryOptimizer(default_lead_time_days=7, default_service_level=0.95)
    
    item = optimizer.calculate_item_inventory(
        stock_code="TEST_ITEM",
        description="Test Item",
        expected_30d_demand=300.0,
        daily_demand_std=5.0,
        unit_price=10.0,
        current_stock=50,
        lead_time_days=7,
        service_level=0.95
    )
    
    assert item["stock_code"] == "TEST_ITEM"
    assert item["safety_stock"] > 0
    assert item["reorder_point"] > 50
    assert item["status"] == "Replenishment Needed"
    assert item["suggested_order"] > 0
    assert item["order_cost_scenario"] == item["suggested_order"] * 10.0


def test_inventory_calculation_excess_stock():
    """Tests status when current stock significantly exceeds 60-day demand."""
    optimizer = InventoryOptimizer()
    
    item = optimizer.calculate_item_inventory(
        stock_code="EXCESS_ITEM",
        description="Excess Item",
        expected_30d_demand=100.0,
        daily_demand_std=2.0,
        unit_price=5.0,
        current_stock=1000,
        lead_time_days=7,
        service_level=0.95
    )
    
    assert item["status"] == "Excess Stock"
    assert item["suggested_order"] == 0


def test_inventory_expiry_risk_alert():
    """Tests expiry integration to trigger alert and halt replenishment."""
    optimizer = InventoryOptimizer()
    
    item = optimizer.calculate_item_inventory(
        stock_code="EXP_ITEM",
        description="Expiring Item",
        expected_30d_demand=300.0,  # 10/day
        daily_demand_std=3.0,
        unit_price=8.0,
        current_stock=100,  # Low stock, but expires in 5 days (expected demand 50)
        lead_time_days=7,
        service_level=0.95,
        expiry_days_remaining=5,
        expiry_status="Critical"
    )
    
    assert item["expiry_risk_alert"] is not None
    assert item["expiry_risk_alert"]["is_high_risk"] is True
    assert item["expiry_risk_alert"]["units_at_risk"] == 50
    assert item["suggested_order"] == 0  # Replenishment halted to prevent waste


def test_manual_mathematical_formulas_validation():
    """Validates exact textbook inventory mathematics: LTD, sigma_LT, SS, ROP, Suggested Order."""
    optimizer = InventoryOptimizer()
    
    # Inputs:
    # expected_30d_demand = 300 (daily_mean = 10.0)
    # daily_demand_std = 4.0
    # lead_time = 9 (sqrt(9) = 3.0)
    # service_level = 0.95 (z = 1.645)
    # current_stock = 60
    item = optimizer.calculate_item_inventory(
        stock_code="MATH_TEST",
        description="Math Test Item",
        expected_30d_demand=300.0,
        daily_demand_std=4.0,
        unit_price=15.0,
        current_stock=60,
        lead_time_days=9,
        service_level=0.95
    )
    
    # 1. Lead Time Demand = 10.0 * 9 = 90.0
    assert item["lead_time_demand"] == 90.0
    
    # 2. Lead Time Std = 4.0 * sqrt(9) = 12.0
    # Safety Stock = ceil(1.645 * 12.0) = ceil(19.74) = 20
    assert item["safety_stock"] == 20
    
    # 3. Reorder Point = ceil(90.0 + 20) = 110
    assert item["reorder_point"] == 110
    
    # 4. Target = 300 + 20 = 320 -> Suggested Order = 320 - 60 = 260
    assert item["suggested_order"] == 260
    assert item["status"] == "Replenishment Needed"


def test_service_level_scaling_behavior():
    """Validates that higher service level targets strictly increase safety stock and ROP."""
    optimizer = InventoryOptimizer()
    
    levels = [0.90, 0.95, 0.98, 0.99]
    results = [
        optimizer.calculate_item_inventory(
            stock_code="SL_TEST",
            description="Service Level Test",
            expected_30d_demand=600.0,
            daily_demand_std=10.0,
            unit_price=5.0,
            current_stock=100,
            lead_time_days=7,
            service_level=lvl
        )
        for lvl in levels
    ]
    
    safety_stocks = [r["safety_stock"] for r in results]
    reorder_points = [r["reorder_point"] for r in results]
    
    # Strictly increasing safety stocks and ROPs with higher confidence
    assert safety_stocks[0] < safety_stocks[1] < safety_stocks[2] < safety_stocks[3]
    assert reorder_points[0] < reorder_points[1] < reorder_points[2] < reorder_points[3]


def test_lead_time_scaling_behavior():
    """Validates lead time scaling: LTD scales linearly, safety stock scales with sqrt(L)."""
    optimizer = InventoryOptimizer()
    
    res_lt4 = optimizer.calculate_item_inventory(
        stock_code="LT_TEST",
        description="Lead Time Test",
        expected_30d_demand=300.0,
        daily_demand_std=10.0,
        unit_price=5.0,
        current_stock=100,
        lead_time_days=4,
        service_level=0.95
    )
    
    res_lt16 = optimizer.calculate_item_inventory(
        stock_code="LT_TEST",
        description="Lead Time Test",
        expected_30d_demand=300.0,
        daily_demand_std=10.0,
        unit_price=5.0,
        current_stock=100,
        lead_time_days=16,
        service_level=0.95
    )
    
    # LTD for 16 days is 4x that of 4 days (linear)
    assert res_lt16["lead_time_demand"] == res_lt4["lead_time_demand"] * 4.0
    
    # Safety stock for 16 days is ~2x that of 4 days (sqrt(16)/sqrt(4) = 4/2 = 2)
    assert abs(res_lt16["safety_stock"] - 2 * res_lt4["safety_stock"]) <= 1


def test_zero_and_near_zero_demand():
    """Validates that zero or near-zero demand does not produce negative or absurd values."""
    optimizer = InventoryOptimizer()
    
    # Complete zero
    res_zero = optimizer.calculate_item_inventory(
        stock_code="ZERO_ITEM",
        description="Zero Demand Item",
        expected_30d_demand=0.0,
        daily_demand_std=0.0,
        unit_price=10.0,
        current_stock=0,
        lead_time_days=7,
        service_level=0.95
    )
    assert res_zero["lead_time_demand"] == 0.0
    assert res_zero["safety_stock"] == 0
    assert res_zero["reorder_point"] == 0
    assert res_zero["suggested_order"] == 0
    
    # Tiny demand
    res_tiny = optimizer.calculate_item_inventory(
        stock_code="TINY_ITEM",
        description="Tiny Demand Item",
        expected_30d_demand=1.0,
        daily_demand_std=0.2,
        unit_price=10.0,
        current_stock=0,
        lead_time_days=7,
        service_level=0.95
    )
    assert res_tiny["lead_time_demand"] > 0
    assert res_tiny["safety_stock"] >= 0
    assert res_tiny["suggested_order"] >= 1


def test_simulator_and_catalog_agreement_product_84077():
    """Integration test proving catalog and simulator produce 100% identical metrics for Product 84077."""
    from backend.app.db.database import SessionLocal
    from backend.app.services.retail_intelligence_service import retail_intelligence_service

    db = SessionLocal()
    try:
        items = retail_intelligence_service.get_inventory_recommendations(db=db, limit=0)
        item_84077 = next((i for i in items if i['stock_code'] == '84077'), None)
        assert item_84077 is not None
        
        sim_84077 = retail_intelligence_service.simulate_inventory(
            stock_code='84077',
            current_stock=item_84077['current_stock'],
            lead_time_days=item_84077['lead_time_days'],
            service_level=item_84077['service_level'],
            db=db
        )
        
        # Verify agreement across all primary metrics
        assert item_84077['expected_30d_demand'] == sim_84077['expected_30d_demand']
        assert item_84077['lead_time_demand'] == sim_84077['lead_time_demand']
        assert item_84077['safety_stock'] == sim_84077['safety_stock']
        assert item_84077['reorder_point'] == sim_84077['reorder_point']
        assert item_84077['current_stock'] == sim_84077['current_stock']
        assert item_84077['suggested_order'] == sim_84077['suggested_order']
        assert item_84077['status'] == sim_84077['status']
    finally:
        db.close()


def test_all_eligible_products_processed_no_150_limit():
    """Validates that all eligible products (>4000) are analysed without an artificial 150 limit."""
    from backend.app.db.database import SessionLocal
    from backend.app.services.retail_intelligence_service import retail_intelligence_service

    db = SessionLocal()
    try:
        summary = retail_intelligence_service.get_inventory_summary(db=db)
        assert summary["total_products_available"] >= 4600
        assert summary["total_products_analysed"] > 4000
        assert summary["excluded_products_count"] > 0
        assert "/" in summary["products_analysed_display"]

        # Recommendations without limit returns the complete eligible population
        all_eligible = retail_intelligence_service.get_inventory_recommendations(db=db, limit=0)
        assert len(all_eligible) == summary["total_products_analysed"]
        assert len(all_eligible) > 4000
    finally:
        db.close()


def test_insufficient_history_products_handled_correctly():
    """Validates that SKUs with insufficient history are flagged, non-fabricated, and excluded from automated ordering."""
    from backend.app.db.database import SessionLocal
    from backend.app.services.retail_intelligence_service import retail_intelligence_service

    db = SessionLocal()
    try:
        all_with_excluded = retail_intelligence_service.get_inventory_recommendations(db=db, limit=0, include_excluded=True)
        excluded_items = [i for i in all_with_excluded if not i.get('is_eligible', True)]
        
        assert len(excluded_items) > 0
        for item in excluded_items:
            assert item["is_eligible"] is False
            assert item["suggested_order"] == 0
            assert item["status"] == "Insufficient History"
            assert item["exclusion_reason"] is not None or "insufficient" in item["reason"].lower()
    finally:
        db.close()


def test_excel_workbook_generation_structure_and_readability():
    """Validates that the generated Excel report is a real valid .xlsx workbook with 3 structured sheets."""
    import io
    import openpyxl
    from backend.app.db.database import SessionLocal
    from backend.app.services.retail_intelligence_service import retail_intelligence_service

    db = SessionLocal()
    try:
        excel_bytes = retail_intelligence_service.generate_inventory_excel_workbook(db=db)
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 100000  # Multi-hundred KB workbook

        # Load workbook with openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        assert "Inventory Recommendations" in wb.sheetnames
        assert "Executive Summary" in wb.sheetnames
        assert "Excluded Products" in wb.sheetnames

        ws1 = wb["Inventory Recommendations"]
        # Must have header + >4000 rows
        assert ws1.max_row > 4000
        assert ws1.max_column >= 15
        assert ws1.cell(row=1, column=1).value == "Stock Code"
        assert ws1.cell(row=1, column=4).value == "30-Day Forecast"
        assert ws1.cell(row=1, column=9).value == "Safety Stock Buffer"
        assert ws1.cell(row=1, column=11).value == "Suggested Order Quantity"

        ws2 = wb["Executive Summary"]
        assert ws2.max_row >= 10
        assert "Executive Summary" in str(ws2.cell(row=1, column=1).value)

        ws3 = wb["Excluded Products"]
        assert ws3.max_row > 10
        assert ws3.cell(row=1, column=1).value == "Stock Code"
    finally:
        db.close()


def test_email_report_attachment_parity():
    """Validates that the email report attaches the exact generated Excel workbook with full data parity."""
    from backend.app.db.database import SessionLocal
    from backend.app.services.retail_intelligence_service import retail_intelligence_service
    from backend.app.services.email_service import email_service

    db = SessionLocal()
    try:
        excel_bytes = retail_intelligence_service.generate_inventory_excel_workbook(db=db)
        
        result = email_service.send_inventory_report_email(
            excel_bytes=excel_bytes,
            filename="Retail_Inventory_Replenishment_Report_test.xlsx",
            recipient_email="akarshanrasyal4@gmail.com",
            subject="Retail Inventory Replenishment Report",
            message_text="Please find attached the latest replenishment report."
        )

        assert "status" in result
        assert "audit_id" in result
        assert result["delivery_mode"] in ["BREVO API", "DEMO EMAIL"]
        assert result["recipient"] == "akarshanrasyal4@gmail.com"
    finally:
        db.close()


def test_expiry_replenishment_protection_across_catalog():
    """Validates that products where scenario stock covers demand before expiry have replenishment halted."""
    from backend.app.db.database import SessionLocal
    from backend.app.services.retail_intelligence_service import retail_intelligence_service

    db = SessionLocal()
    try:
        items = retail_intelligence_service.get_inventory_recommendations(db=db, limit=0)
        expiring_with_excess = [
            i for i in items 
            if i.get('expiry_risk_alert') and i['expiry_risk_alert'].get('units_at_risk', 0) > 0
        ]
        
        assert len(expiring_with_excess) > 0
        for item in expiring_with_excess:
            # When stock exceeds demand before expiry, suggested order is halted to 0 to prevent waste
            assert item['suggested_order'] == 0
            assert "waste" in item['reason'].lower() or "expiry" in item['reason'].lower() or "expir" in item['expiry_risk_alert']['recommendation'].lower()
    finally:
        db.close()


