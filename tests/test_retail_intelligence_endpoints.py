"""
Integration tests for the new Retail Intelligence REST Endpoints.
"""
import pytest
from fastapi.testclient import TestClient
from backend.app.main import app

client = TestClient(app)


def test_forecasting_summary_endpoint():
    """Tests GET /api/forecasting/summary."""
    response = client.get("/api/forecasting/summary?dashboard_id=default")
    assert response.status_code == 200
    data = response.json()
    assert "products_forecasted" in data
    assert "total_expected_30d_units" in data
    assert "avg_mae" in data
    assert "products_rising_demand" in data


def test_forecasting_products_endpoint():
    """Tests GET /api/forecasting/products."""
    response = client.get("/api/forecasting/products?dashboard_id=default&limit=10")
    assert response.status_code == 200
    data = response.json()
    assert isinstance(data, list)
    assert len(data) > 0
    first = data[0]
    assert "stock_code" in first
    assert "expected_30d_demand" in first
    assert "recent_30d_demand" in first
    assert "trend_pct" in first
    assert "trend_direction" in first
    assert first["trend_direction"] in ["Rising", "Falling", "Stable"]


def test_forecasting_product_detail_endpoint():
    """Tests GET /api/forecasting/product/{stock_code}."""
    response = client.get("/api/forecasting/product/85123A?dashboard_id=default")
    assert response.status_code == 200
    data = response.json()
    assert data["stock_code"] == "85123A"
    assert "history" in data
    assert "forecast" in data
    assert len(data["forecast"]) == 30
    assert "trend_pct" in data
    assert "trend_direction" in data
    assert data["trend_direction"] in ["Rising", "Falling", "Stable"]
    first_f = data["forecast"][0]
    assert "forecast_units" in first_f
    assert "lower_bound" in first_f
    assert "upper_bound" in first_f


def test_inventory_summary_endpoint():
    """Tests GET /api/inventory/summary."""
    response = client.get("/api/inventory/summary?dashboard_id=default")
    assert response.status_code == 200
    data = response.json()
    assert "total_products_analysed" in data
    assert "replenishment_needed_count" in data
    assert "total_suggested_order_units" in data


def test_inventory_simulate_endpoint():
    """Tests POST /api/inventory/simulate."""
    payload = {
        "stock_code": "85123A",
        "current_stock": 150,
        "lead_time_days": 10,
        "service_level": 0.95,
        "holding_cost_pct": 0.20
    }
    response = client.post("/api/inventory/simulate?dashboard_id=default", json=payload)
    assert response.status_code == 200
    data = response.json()
    assert data["stock_code"] == "85123A"
    assert "safety_stock" in data
    assert "reorder_point" in data
    assert "suggested_order" in data


def test_pricing_summary_endpoint():
    """Tests GET /api/pricing/summary."""
    response = client.get("/api/pricing/summary?dashboard_id=default")
    assert response.status_code == 200
    data = response.json()
    assert "total_products_analysed" in data
    assert "elastic_products_count" in data
    assert "inelastic_products_count" in data
    assert data["total_products_analysed"] > 0


def test_pricing_products_endpoint():
    """Tests GET /api/pricing/products returns full population with search and filters, without artificial limits."""
    # 1. Complete catalog list without limit
    resp = client.get("/api/pricing/products?dashboard_id=default")
    assert resp.status_code == 200
    prods = resp.json()
    assert len(prods) == 4631  # Complete product population
    
    # 2. Check first and end-of-population products
    first = prods[0]
    last = prods[-1]
    assert "stock_code" in first
    assert "avg_price" in first
    assert first["data_provenance"] == "Real historical transactions"
    assert "stock_code" in last

    # 3. Search by StockCode for product far beyond index 150 (near the end of catalog)
    resp_search_code = client.get("/api/pricing/products?dashboard_id=default&search=35999")
    assert resp_search_code.status_code == 200
    code_results = resp_search_code.json()
    assert len(code_results) >= 1
    assert any(p["stock_code"] == "35999" for p in code_results)
    p_end = next(p for p in code_results if p["stock_code"] == "35999")
    assert p_end["description"] == "S/6 Scandinavian Heart T-Light"

    # 4. Search by Description across full catalog
    resp_search_desc = client.get("/api/pricing/products?dashboard_id=default&search=SCANDINAVIAN")
    assert resp_search_desc.status_code == 200
    desc_results = resp_search_desc.json()
    assert len(desc_results) > 0
    assert any("SCANDINAVIAN" in p["description"].upper() for p in desc_results)

    # 5. Verify optimiser and simulator run successfully on this non-top-150 item
    resp_opt_end = client.post("/api/pricing/optimize?dashboard_id=default", json={
        "stock_code": "35999",
        "objective": "revenue"
    })
    assert resp_opt_end.status_code == 200
    opt_end_data = resp_opt_end.json()
    assert opt_end_data["stock_code"] == "35999"
    assert opt_end_data["historical_avg_price"] == 2.55


def test_pricing_simulate_endpoint_with_and_without_unit_cost():
    """Tests POST /api/pricing/simulate with optional unit cost and verifies no fabrication."""
    # Scenario without unit cost
    payload_no_cost = {
        "stock_code": "85123A",
        "price_change_pct": -10.0,
        "scenario_unit_cost": None
    }
    resp1 = client.post("/api/pricing/simulate?dashboard_id=default", json=payload_no_cost)
    assert resp1.status_code == 200
    data1 = resp1.json()
    assert data1["stock_code"] == "85123A"
    assert "new_price" in data1
    assert "expected_quantity" in data1
    assert "revenue_difference" in data1
    assert data1["scenario_unit_cost"] is None
    assert data1["scenario_profit"] is None
    assert data1["profit_difference"] is None

    # Scenario with explicit user unit cost assumption
    payload_with_cost = {
        "stock_code": "85123A",
        "price_change_pct": -10.0,
        "scenario_unit_cost": 1.50
    }
    resp2 = client.post("/api/pricing/simulate?dashboard_id=default", json=payload_with_cost)
    assert resp2.status_code == 200
    data2 = resp2.json()
    assert data2["scenario_unit_cost"] == 1.50
    assert data2["scenario_profit"] is not None
    assert data2["baseline_profit"] is not None
    assert data2["profit_difference"] is not None
    assert round(data2["scenario_profit"] - data2["baseline_profit"], 2) == data2["profit_difference"]


def test_pricing_optimize_endpoint():
    """Tests POST /api/pricing/optimize for profit and revenue objectives."""
    # 1. Profit objective with unit cost
    payload_profit = {
        "stock_code": "85123A",
        "objective": "profit",
        "unit_cost": 1.20
    }
    resp1 = client.post("/api/pricing/optimize?dashboard_id=default", json=payload_profit)
    assert resp1.status_code == 200
    data1 = resp1.json()
    assert data1["stock_code"] == "85123A"
    assert data1["objective"] == "profit"
    assert "recommended_price" in data1
    assert "expected_30d_quantity" in data1
    assert "expected_30d_revenue" in data1
    assert "expected_30d_profit" in data1
    assert data1["unit_cost"] == 1.20

    # 2. Revenue objective without unit cost
    payload_rev = {
        "stock_code": "85123A",
        "objective": "revenue",
        "unit_cost": None
    }
    resp2 = client.post("/api/pricing/optimize?dashboard_id=default", json=payload_rev)
    assert resp2.status_code == 200
    data2 = resp2.json()
    assert data2["objective"] == "revenue"
    assert data2["unit_cost"] is None
    assert data2["expected_30d_profit"] is None  # Never fabricated


def test_pricing_export_analysis_excel_endpoint():
    """Tests GET /api/pricing/export-analysis generates a valid .xlsx analysis workbook."""
    import io
    import openpyxl

    resp = client.get("/api/pricing/export-analysis?stock_code=85123A&objective=profit&unit_cost=1.20&dashboard_id=default")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert len(resp.content) > 1000

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Pricing & Profit Decision" in wb.sheetnames
    ws = wb["Pricing & Profit Decision"]
    assert "Pricing & Profit Optimisation Analysis" in str(ws["A1"].value)


def test_monitoring_summary_endpoint():
    """Tests GET /api/monitoring/summary."""
    response = client.get("/api/monitoring/summary?dashboard_id=default")
    assert response.status_code == 200
    data = response.json()
    assert "overall_system_health" in data
    assert "feature_drift_results" in data
    assert "demand_alerts" in data


def test_csv_download_endpoints():
    """Tests the new CSV export download endpoints."""
    # Forecast CSV
    fc_resp = client.get("/api/forecasting/download?dashboard_id=default")
    assert fc_resp.status_code == 200
    assert "text/csv" in fc_resp.headers["content-type"]
    assert "stock_code" in fc_resp.text

    # Inventory CSV
    inv_resp = client.get("/api/inventory/download?dashboard_id=default")
    assert inv_resp.status_code == 200
    assert "text/csv" in inv_resp.headers["content-type"]
    assert "reorder_point" in inv_resp.text

    # Pricing CSV
    pr_resp = client.get("/api/pricing/download?dashboard_id=default")
    assert pr_resp.status_code == 200
    assert "text/csv" in pr_resp.headers["content-type"]
    assert "elasticity" in pr_resp.text

    # Monitoring CSV
    mon_resp = client.get("/api/monitoring/download?dashboard_id=default")
    assert mon_resp.status_code == 200
    assert "text/csv" in mon_resp.headers["content-type"]
    assert "feature_name" in mon_resp.text


def test_inventory_excel_export_endpoint():
    """Tests GET /api/inventory/export-excel generates a valid .xlsx workbook."""
    import io
    import openpyxl

    resp = client.get("/api/inventory/export-excel?dashboard_id=default")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert len(resp.content) > 100000

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Inventory Recommendations" in wb.sheetnames
    assert "Executive Summary" in wb.sheetnames
    assert "Excluded Products" in wb.sheetnames


def test_inventory_email_report_endpoint(monkeypatch):
    """Tests POST /api/inventory/email-report attaches the Excel workbook (mocked to protect email limits)."""
    from backend.app.services.email_service import email_service
    monkeypatch.setattr(email_service, "send_inventory_report_email", lambda **kwargs: {"status": "success", "audit_id": "mocked-test-audit-id", "recipient": kwargs.get("recipient_email")})

    payload = {
        "recipient_email": "test_recipient@example.com",
        "subject": "Retail Inventory Replenishment Report",
        "message": "Please find attached the latest report."
    }
    resp = client.post("/api/inventory/email-report?dashboard_id=default", json=payload)
    assert resp.status_code == 200
    data = resp.json()
    assert "status" in data
    assert "audit_id" in data
